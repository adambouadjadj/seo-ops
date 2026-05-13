import re
from pathlib import Path
from datetime import date, timedelta

PROJECT_ROOT = Path(__file__).resolve().parents[3]  # seo-ops/
INPUTS_DIR = PROJECT_ROOT / 'inputs'
TASKS_DIR = PROJECT_ROOT / 'tasks'
OUTPUT_DIR = PROJECT_ROOT / 'output' / 'reporting'


def fetch_gsc_for_reporting() -> Path:
    """
    Récupère les données GSC (7j N vs 7j N-1) via l'API et génère le fichier
    'GSC performance AB DD-MM-YY.xlsx' dans inputs/.
    Retourne le Path du fichier généré.
    """
    import sys
    sys.path.insert(0, str(PROJECT_ROOT / 'tools'))
    from gsc_client import search_analytics

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError("openpyxl manquant — pip install openpyxl")

    # Plages de dates (lag GSC = 3 jours)
    end_n     = date.today() - timedelta(days=3)
    start_n   = end_n - timedelta(days=6)
    end_prev  = start_n - timedelta(days=1)
    start_prev = end_prev - timedelta(days=6)

    fmt = lambda d: d.strftime("%Y-%m-%d")

    # Requêtes API
    q_n    = search_analytics(fmt(start_n),    fmt(end_n),    ["query"], row_limit=1000)
    q_prev = search_analytics(fmt(start_prev), fmt(end_prev), ["query"], row_limit=1000)
    p_n    = search_analytics(fmt(start_n),    fmt(end_n),    ["page"],  row_limit=1000)
    p_prev = search_analytics(fmt(start_prev), fmt(end_prev), ["page"],  row_limit=1000)

    def to_dict(rows, key_index=0):
        return {
            r["keys"][key_index]: {
                "clicks":      int(r.get("clicks", 0)),
                "impressions": int(r.get("impressions", 0)),
                "ctr":         round(r.get("ctr", 0), 4),
                "position":    round(r.get("position", 0), 1),
            }
            for r in rows
        }

    qn_map    = to_dict(q_n)
    qprev_map = to_dict(q_prev)
    pn_map    = to_dict(p_n)
    pprev_map = to_dict(p_prev)

    HEADERS = [
        "Top queries", "Last 7 days Clicks", "Previous 7 days Clicks",
        "Last 7 days Impressions", "Previous 7 days Impressions",
        "Last 7 days CTR", "Previous 7 days CTR",
        "Last 7 days Position", "Previous 7 days Position",
    ]

    HDR_FILL = PatternFill("solid", fgColor="1A3A5C")
    HDR_FONT = Font(bold=True, color="FFFFFF")

    def write_sheet(ws, current_map, prev_map, label_col):
        ws.cell(1, 1).value = label_col
        for i, h in enumerate(HEADERS[1:], 2):
            ws.cell(1, i).value = h
        for col in range(1, len(HEADERS) + 1):
            c = ws.cell(1, col)
            c.font, c.fill = HDR_FONT, HDR_FILL
            c.alignment = Alignment(horizontal="center")
        ws.column_dimensions["A"].width = 55
        for col_letter in ["B", "C", "D", "E", "F", "G", "H", "I"]:
            ws.column_dimensions[col_letter].width = 22

        # Trier par impressions N desc
        keys = sorted(current_map.keys(), key=lambda k: -current_map[k]["impressions"])
        for row_i, key in enumerate(keys, 2):
            n    = current_map[key]
            prev = prev_map.get(key, {"clicks": 0, "impressions": 0, "ctr": 0.0, "position": 0.0})
            ws.cell(row_i, 1, key)
            ws.cell(row_i, 2, n["clicks"])
            ws.cell(row_i, 3, prev["clicks"])
            ws.cell(row_i, 4, n["impressions"])
            ws.cell(row_i, 5, prev["impressions"])
            ws.cell(row_i, 6, n["ctr"])
            ws.cell(row_i, 7, prev["ctr"])
            ws.cell(row_i, 8, n["position"])
            ws.cell(row_i, 9, prev["position"])

    wb = openpyxl.Workbook()
    ws_queries = wb.active
    ws_queries.title = "Queries"
    write_sheet(ws_queries, qn_map, qprev_map, "Top queries")

    ws_pages = wb.create_sheet("Pages")
    write_sheet(ws_pages, pn_map, pprev_map, "Top pages")

    INPUTS_DIR.mkdir(exist_ok=True)
    date_str = end_n.strftime("%d-%m-%y")
    out_path = INPUTS_DIR / f"GSC performance AB {date_str}.xlsx"
    wb.save(str(out_path))
    return out_path


def _current_friday() -> date:
    today = date.today()
    days_until_friday = (4 - today.weekday()) % 7
    return today + timedelta(days=days_until_friday)


def _extract_date_from_filename(filename: str) -> str | None:
    """Try to extract a DD-MM-YY date pattern from a filename."""
    m = re.search(r'(\d{2})[-_](\d{2})[-_](\d{2,4})', filename)
    if m:
        return f'{m.group(1)}-{m.group(2)}-{m.group(3)[-2:]}'
    return None


def stage_files(monitorank_file, gsc_file, task_filename: str, gsc_prefetched: str = '') -> dict:
    """Save uploaded files to inputs/ using the standard naming convention.

    Naming: 'Suivi positionnement AB DD-MM-YY.xlsx' and 'GSC performance AB DD-MM-YY.xlsx'.
    The date is extracted from the uploaded filename or derived from the current Friday.
    Returns {'monitorank': Path, 'gsc': Path, 'task': Path|None, 'week_label': str}.
    """
    INPUTS_DIR.mkdir(exist_ok=True)

    raw_date = _extract_date_from_filename(monitorank_file.filename or '')
    if not raw_date and gsc_file:
        raw_date = _extract_date_from_filename(gsc_file.filename or '')
    if not raw_date and gsc_prefetched:
        raw_date = _extract_date_from_filename(gsc_prefetched)
    if not raw_date:
        raw_date = _current_friday().strftime('%d-%m-%y')

    mono_name = f'Suivi positionnement AB {raw_date}.xlsx'
    mono_path = INPUTS_DIR / mono_name
    monitorank_file.save(str(mono_path))

    if gsc_prefetched:
        gsc_path = INPUTS_DIR / gsc_prefetched
    else:
        gsc_name = f'GSC performance AB {raw_date}.xlsx'
        gsc_path = INPUTS_DIR / gsc_name
        gsc_file.save(str(gsc_path))

    task_path = (TASKS_DIR / task_filename) if task_filename else None

    return {
        'monitorank': mono_path,
        'gsc': gsc_path,
        'task': task_path,
        'week_label': raw_date,
    }


def get_available_task_files() -> list:
    """List tasks/*.txt sorted by modification time desc.

    Returns [(filename, label)] where label is a human-readable date.
    """
    if not TASKS_DIR.exists():
        return []
    files = sorted(
        TASKS_DIR.glob('*.txt'),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    result = []
    for f in files:
        # task_16_02_26.txt → 16/02/26
        label = f.stem.replace('task_', '').replace('_', '/')
        result.append((f.name, label))
    return result


def get_past_reports() -> list:
    """
    List paired reporting HTML + MD files sorted by date desc.
    Handles both new (subfolder DD-MM-YY/) and old (flat) structures.
    Returns paths relative to OUTPUT_DIR.
    """
    if not OUTPUT_DIR.exists():
        return []

    # Collect HTML files: nested (new) + flat (old), deduplicate by stem key
    by_key: dict = {}
    for f in OUTPUT_DIR.glob('*/reporting_*_email.html'):
        key = f.name.replace('_email.html', '')
        by_key[key] = f
    for f in OUTPUT_DIR.glob('reporting_*_email.html'):
        key = f.name.replace('_email.html', '')
        if key not in by_key:
            by_key[key] = f

    reports = []
    for key, html_f in by_key.items():
        md_f = html_f.with_name(key + '.md')
        m = re.search(r'(\d{2}-\d{2}-\d{2,4})$', key)
        date_str = m.group(1) if m else key
        reports.append({
            'name': key,
            'html_path': str(html_f.relative_to(OUTPUT_DIR)),
            'md_path': str(md_f.relative_to(OUTPUT_DIR)) if md_f.exists() else None,
            'date': date_str,
        })

    def _parse_date(date_str):
        try:
            return date(*reversed([int(x) for x in date_str.split('-')]))
        except Exception:
            return date.min

    reports.sort(key=lambda r: _parse_date(r['date']), reverse=True)
    return reports
