#!/usr/bin/env python3
"""
backlink_check.py — Verification des domaines prioritaires pour l'outreach

Pour chaque domaine HAUTE priorite du fichier backlink_gap enrichi :
- Verifie le statut HTTP (actif / mort / redirige)
- Detecte la langue du site (FR / autre)
- Score spam : analyse title, H1, nb liens externes, mots-cles suspects

Usage:
    python tools/backlink_check.py
    python tools/backlink_check.py --input output/backlinks/mon_fichier_enriched.xlsx
    python tools/backlink_check.py --priority HAUTE MOYENNE
"""

import argparse
import re
import time
import urllib.request
import urllib.error
from html.parser import HTMLParser
from pathlib import Path

# ── Config ────────────────────────────────────────────────────────────────────

PROJECT_ROOT = Path(__file__).parent.parent
OUTPUT_DIR   = PROJECT_ROOT / "output" / "backlinks"

PAUSE        = 1.0   # secondes entre requetes
TIMEOUT      = 8     # timeout par domaine
MAX_BYTES    = 80_000  # on lit max 80KB de la homepage

SPAM_KEYWORDS = [
    "casino", "poker", "slot", "betting", "bet ", "paris sportif",
    "viagra", "cialis", "pharmacy", "pharmacie", "medication", "pill",
    "forex", "crypto", "bitcoin", "trading", "binary option",
    "escort", "xxx", "porn", "adult",
    "payday loan", "credit rapide", "pret sans",
    "replica", "fake", "counterfeit",
    "seo links", "buy links", "backlinks",
]

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept-Language": "fr-FR,fr;q=0.9,en;q=0.8",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}


# ── HTML Parser ───────────────────────────────────────────────────────────────

class HomepageParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.title       = ""
        self.h1s         = []
        self.lang        = ""
        self.ext_links   = 0
        self.meta_lang   = ""
        self._in_title   = False
        self._in_h1      = False
        self._current_h1 = ""

    def handle_starttag(self, tag, attrs):
        attrs_dict = dict(attrs)

        if tag == "html":
            self.lang = attrs_dict.get("lang", "")[:5]

        elif tag == "meta":
            http_equiv = attrs_dict.get("http-equiv", "").lower()
            name       = attrs_dict.get("name", "").lower()
            content    = attrs_dict.get("content", "")
            if http_equiv == "content-language" or name == "language":
                self.meta_lang = content[:5]

        elif tag == "title":
            self._in_title = True

        elif tag == "h1":
            self._in_h1      = True
            self._current_h1 = ""

        elif tag == "a":
            href = attrs_dict.get("href", "")
            if href.startswith("http") and not href.startswith("https://www.") and \
               not href.startswith("http://www."):
                self.ext_links += 1

    def handle_endtag(self, tag):
        if tag == "title":
            self._in_title = False
        elif tag == "h1":
            if self._current_h1.strip():
                self.h1s.append(self._current_h1.strip()[:120])
            self._in_h1      = False
            self._current_h1 = ""

    def handle_data(self, data):
        if self._in_title and not self.title:
            self.title += data.strip()
        if self._in_h1:
            self._current_h1 += data


# ── Checker ───────────────────────────────────────────────────────────────────

def check_domain(domain: str) -> dict:
    """
    Retourne un dict avec statut, langue, nb_ext_links, title, h1, spam_score, spam_flags.
    """
    result = {
        "statut":       "erreur",
        "code_http":    0,
        "url_finale":   "",
        "langue":       "",
        "title":        "",
        "h1":           "",
        "nb_ext_links": 0,
        "spam_score":   0,
        "spam_flags":   "",
    }

    for scheme in ("https", "http"):
        url = f"{scheme}://{domain}"
        try:
            req  = urllib.request.Request(url, headers=HEADERS, method="GET")
            resp = urllib.request.urlopen(req, timeout=TIMEOUT)

            result["code_http"]  = resp.status
            result["url_finale"] = resp.url
            result["statut"]     = "actif"

            # Detecter redirections vers autre domaine
            final_host = re.sub(r"https?://(www\.)?", "", resp.url).split("/")[0]
            original   = domain.lstrip("www.")
            if final_host and final_host != original and not final_host.endswith("." + original):
                result["statut"] = f"redirige -> {final_host}"

            # Lire le HTML
            raw = resp.read(MAX_BYTES)
            charset = "utf-8"
            ct = resp.headers.get("Content-Type", "")
            m  = re.search(r"charset=([^\s;]+)", ct, re.I)
            if m:
                charset = m.group(1).strip().lower()
            try:
                html = raw.decode(charset, errors="replace")
            except (LookupError, UnicodeDecodeError):
                html = raw.decode("utf-8", errors="replace")

            # Parser
            parser = HomepageParser()
            try:
                parser.feed(html)
            except Exception:
                pass

            result["title"]        = parser.title[:120]
            result["h1"]           = parser.h1s[0] if parser.h1s else ""
            result["nb_ext_links"] = parser.ext_links

            # Langue
            lang_raw = parser.lang or parser.meta_lang or \
                       resp.headers.get("Content-Language", "")
            result["langue"] = detect_lang(lang_raw, html)

            # Score spam
            spam_score, spam_flags = compute_spam_score(
                result["title"], result["h1"], result["nb_ext_links"], html
            )
            result["spam_score"] = spam_score
            result["spam_flags"] = spam_flags
            break

        except urllib.error.HTTPError as e:
            result["code_http"] = e.code
            result["statut"]    = f"HTTP {e.code}"
            break
        except urllib.error.URLError:
            result["statut"] = "mort"
            # On essaie http si https a echoue
            if scheme == "https":
                continue
            break
        except Exception as e:
            result["statut"] = f"erreur: {type(e).__name__}"
            break

    return result


def detect_lang(lang_raw: str, html: str) -> str:
    """Retourne 'fr', 'en', autre code, ou '' si inconnu."""
    if lang_raw:
        code = lang_raw.lower().split("-")[0].split("_")[0]
        if len(code) == 2:
            return code

    # Heuristique : compter mots FR vs EN dans le HTML visible
    text = re.sub(r"<[^>]+>", " ", html).lower()
    fr_words = len(re.findall(
        r"\b(le|la|les|de|du|des|et|en|pour|dans|avec|sur|croisiere|voyage|mer)\b", text
    ))
    en_words = len(re.findall(
        r"\b(the|and|for|with|cruise|travel|ship|sea|book|view)\b", text
    ))
    if fr_words > en_words * 1.5:
        return "fr"
    if en_words > fr_words * 1.5:
        return "en"
    return ""


def compute_spam_score(title: str, h1: str, nb_ext_links: int, html: str) -> tuple[int, str]:
    """
    Score 0-3 :
      +1 si mot spam dans title ou H1
      +1 si nb liens externes > 80 sur la homepage
      +1 si mot spam dans le body HTML
    """
    score = 0
    flags = []

    combined_visible = (title + " " + h1).lower()
    found_title = [kw for kw in SPAM_KEYWORDS if kw in combined_visible]
    if found_title:
        score += 1
        flags.append(f"title/H1: {', '.join(found_title[:3])}")

    if nb_ext_links > 80:
        score += 1
        flags.append(f"{nb_ext_links} liens ext")

    html_lower = html.lower()
    found_body = [kw for kw in SPAM_KEYWORDS if kw in html_lower]
    # On ignore les mots deja trouves dans title/H1
    found_body_new = [kw for kw in found_body if kw not in found_title]
    if len(found_body_new) >= 3:
        score += 1
        flags.append(f"body: {', '.join(found_body_new[:3])}")

    return score, " | ".join(flags)


# ── Excel ─────────────────────────────────────────────────────────────────────

def load_domains(input_path: Path, priorities: list[str]) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        print("openpyxl manquant -- pip install openpyxl")
        return []

    wb = openpyxl.load_workbook(input_path)
    domains = []
    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("Gap"):
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            priorite = (row[5] or "").upper()
            if priorite in priorities:
                domains.append({
                    "domain":         row[0],
                    "nb_concurrents": row[1],
                    "bas":            row[2],
                    "linkent":        row[3],
                    "ancres":         row[4],
                    "priorite":       priorite,
                    "categorie":      row[6] if len(row) > 6 else "",
                    "sheet":          sheet_name,
                })
    return domains


def generate_report(results: list[dict], output_path: Path):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("openpyxl manquant")
        return

    FILL_STATUT = {
        "actif":   PatternFill("solid", fgColor="D1FAE5"),
        "mort":    PatternFill("solid", fgColor="FEE2E2"),
        "redirige": PatternFill("solid", fgColor="FEF3C7"),
    }
    FILL_LANG = {
        "fr":  PatternFill("solid", fgColor="DBEAFE"),
        "en":  PatternFill("solid", fgColor="FEF9C3"),
    }
    FILL_SPAM = {
        0: PatternFill("solid", fgColor="D1FAE5"),
        1: PatternFill("solid", fgColor="FEF3C7"),
        2: PatternFill("solid", fgColor="FED7AA"),
        3: PatternFill("solid", fgColor="FEE2E2"),
    }
    HDR_FILL = PatternFill("solid", fgColor="1A3A5C")
    HDR_FONT = Font(bold=True, color="FFFFFF")

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Verification"

    headers = [
        "Domaine", "Priorite", "BAS", "Nb conc.", "Linkent",
        "Statut", "Code HTTP", "URL finale",
        "Langue", "Title homepage", "H1",
        "Liens ext.", "Spam score", "Spam flags",
        "Categorie", "Ancres observees",
    ]
    widths = [38, 10, 6, 8, 50, 20, 10, 45, 8, 55, 55, 10, 10, 50, 25, 50]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill = HDR_FONT, HDR_FILL
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    for r_idx, item in enumerate(results, 2):
        chk = item["check"]
        vals = [
            item["domain"],
            item["priorite"],
            item["bas"],
            item["nb_concurrents"],
            item["linkent"],
            chk["statut"],
            chk["code_http"] or "",
            chk["url_finale"],
            chk["langue"],
            chk["title"],
            chk["h1"],
            chk["nb_ext_links"] or "",
            chk["spam_score"],
            chk["spam_flags"],
            item.get("categorie", ""),
            item.get("ancres", ""),
        ]
        for c_idx, val in enumerate(vals, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)

        # Colorier statut (col 6)
        statut_key = "actif" if chk["statut"] == "actif" else \
                     "mort"  if chk["statut"] in ("mort", "erreur") else "redirige"
        ws.cell(r_idx, 6).fill = FILL_STATUT.get(statut_key, FILL_STATUT["redirige"])

        # Colorier langue (col 9)
        if chk["langue"] in FILL_LANG:
            ws.cell(r_idx, 9).fill = FILL_LANG[chk["langue"]]

        # Colorier spam score (col 13)
        ws.cell(r_idx, 13).fill = FILL_SPAM.get(min(chk["spam_score"], 3), FILL_SPAM[0])

    # Feuille recap
    ws2 = wb.create_sheet("Recap")
    actifs    = sum(1 for r in results if r["check"]["statut"] == "actif")
    morts     = sum(1 for r in results if r["check"]["statut"] in ("mort", "erreur"))
    fr        = sum(1 for r in results if r["check"]["langue"] == "fr")
    spam_ok   = sum(1 for r in results if r["check"]["spam_score"] == 0)
    spam_warn = sum(1 for r in results if r["check"]["spam_score"] == 1)
    spam_ko   = sum(1 for r in results if r["check"]["spam_score"] >= 2)

    ws2.append(["Métrique", "Valeur"])
    ws2.append(["Total analysés", len(results)])
    ws2.append(["Actifs", actifs])
    ws2.append(["Morts / erreur", morts])
    ws2.append(["Langue FR", fr])
    ws2.append(["Spam score 0 (clean)", spam_ok])
    ws2.append(["Spam score 1 (attention)", spam_warn])
    ws2.append(["Spam score 2-3 (suspect)", spam_ko])
    ws2.append([])
    ws2.append(["Recommandés outreach", f"{sum(1 for r in results if r['check']['statut'] == 'actif' and r['check']['langue'] == 'fr' and r['check']['spam_score'] == 0)} domaines actifs + FR + clean"])

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 40
    ws2.cell(1, 1).font = ws2.cell(1, 2).font = Font(bold=True)

    wb.save(output_path)
    print(f"Rapport : {output_path}")


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Verification domaines backlink gap")
    parser.add_argument("--input", default=None,
                        help="Fichier Excel enrichi (defaut: dernier _enriched.xlsx)")
    parser.add_argument("--priority", nargs="+", default=["HAUTE"],
                        help="Priorites a verifier (defaut: HAUTE)")
    args = parser.parse_args()

    if args.input:
        input_path = Path(args.input)
    else:
        candidates = sorted(
            OUTPUT_DIR.glob("*_enriched.xlsx"),
            key=lambda p: p.stat().st_mtime, reverse=True
        )
        if not candidates:
            print("Aucun fichier *_enriched.xlsx dans output/backlinks/")
            return
        input_path = candidates[0]

    priorities = [p.upper() for p in args.priority]
    print(f"Input      : {input_path.name}")
    print(f"Priorites  : {', '.join(priorities)}")

    domains = load_domains(input_path, priorities)
    if not domains:
        print("Aucun domaine trouve pour ces priorites.")
        return

    print(f"Domaines   : {len(domains)}\n")

    results = []
    for i, item in enumerate(domains, 1):
        domain = item["domain"]
        print(f"[{i:>3}/{len(domains)}] {domain:<40}", end=" ", flush=True)
        chk = check_domain(domain)
        item["check"] = chk
        results.append(item)

        lang_str  = f"lang={chk['langue'] or '?'}"
        spam_str  = f"spam={chk['spam_score']}"
        print(f"{chk['statut']:<25} {lang_str:<10} {spam_str}")

        time.sleep(PAUSE)

    # Export
    date_str    = __import__("datetime").datetime.now().strftime("%d-%m-%y")
    prio_slug   = "-".join(p.lower() for p in priorities)
    output_path = OUTPUT_DIR / f"backlink_check_{prio_slug}_{date_str}.xlsx"
    generate_report(results, output_path)

    # Recap terminal
    actifs   = sum(1 for r in results if r["check"]["statut"] == "actif")
    fr_clean = sum(1 for r in results if r["check"]["statut"] == "actif"
                   and r["check"]["langue"] == "fr"
                   and r["check"]["spam_score"] == 0)
    print(f"\nActifs      : {actifs}/{len(results)}")
    print(f"FR + clean  : {fr_clean} domaines recommandes pour l'outreach")


if __name__ == "__main__":
    main()
