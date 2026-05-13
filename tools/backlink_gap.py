#!/usr/bin/env python3
"""
backlink_gap.py — Analyse de gap backlinks via API Babbar

Compare les domaines referents d'abcroisiere.com avec ses concurrents
et produit une liste priorisee d'opportunites d'outreach.

Usage:
    python tools/backlink_gap.py
    python tools/backlink_gap.py --include-armateurs
    python tools/backlink_gap.py --url www.abcroisiere.com

Dependances:
    pip install requests openpyxl
"""

import argparse
import json
import time
import urllib.request
import urllib.error
from datetime import datetime
from pathlib import Path

# ── Config ───────────────────────────────────────────────────────────────────

PROJECT_ROOT = Path(__file__).parent.parent
OUTPUT_DIR   = PROJECT_ROOT / "output" / "backlinks"

def load_env():
    env_path = PROJECT_ROOT / "tools" / ".env"
    env = {}
    if env_path.exists():
        for line in env_path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip()
    return env

ENV     = load_env()
API_KEY = ENV.get("BABBAR_API_KEY", "")

BASE_URL = "https://www.babbar.tech/api"
PAUSE    = 2.1  # secondes entre appels API (30 req/min = 2s minimum)

COMPARATEURS = {
    "croisierenet":        "www.croisierenet.com",
    "croisieres.fr":       "www.croisieres.fr",
    "croisieres.com":      "www.croisieres.com",
    "destockagecroisiere": "www.destockagecroisieres.fr",
    "okcroisiere":         "www.okcroisiere.fr",
    "croisiland":          "www.croisiland.com",
}

ARMATEURS = {
    "msc":          "www.msccroisieres.fr",
    "costa":        "www.costacroisieres.fr",
    "croisieurope": "www.croisieurope.com",
}


# ── API ───────────────────────────────────────────────────────────────────────

def _post(endpoint: str, payload: dict) -> dict:
    url  = f"{BASE_URL}/{endpoint}?api_token={API_KEY}"
    data = json.dumps(payload).encode()
    req  = urllib.request.Request(
        url, data=data,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    for attempt in range(10):
        try:
            resp = urllib.request.urlopen(req, timeout=30)
            # Pause proactive si quota presque epuise
            remaining = int(resp.headers.get("x-ratelimit-remaining", 99))
            if remaining <= 2:
                print(f"\n    Quota bas ({remaining} restants) -- pause 62s...", flush=True)
                time.sleep(62)
            return json.loads(resp.read())
        except urllib.error.HTTPError as e:
            if e.code == 429:
                print(f"    Rate limit 429 -- attente 62s... (tentative {attempt+1}/10)")
                time.sleep(62)
            else:
                body = e.read().decode("utf-8", errors="replace")[:200]
                raise RuntimeError(f"HTTP {e.code} sur {endpoint}: {body}") from e
        except Exception:
            if attempt == 9:
                raise
            time.sleep(5)
    raise RuntimeError(f"Rate limit persistant sur {endpoint} apres 10 tentatives")


# ── Fetchers ──────────────────────────────────────────────────────────────────

MAX_PAGES_US   = 250  # couverture quasi-complete pour notre site (~238 pages = 100%)
MAX_PAGES_COMP = 400  # croisierenet/croisieres.fr ont ~175K backlinks -> ~350 pages nécessaires


def fetch_domains_with_bas(host: str, label: str, max_pages: int = MAX_PAGES_COMP) -> dict:
    """
    Recupere les domaines referents + BAS via domain/backlinks/url (BAS desc, offset).
    Retourne {domain: {"bas": int, "anchors": [str]}}.
    """
    from urllib.parse import urlparse

    target = host.lstrip("www.")
    print(f"  {label} ({target})...")
    result  = {}
    offset  = 0
    n_links = 0
    n_total = None

    for page in range(max_pages):
        try:
            data = _post("domain/backlinks/url", {
                "domain": target,
                "limit":  500,
                "sort":   "desc",
                "type":   "babbarAuthorityScore",
                "offset": offset,
            })
        except RuntimeError as e:
            print(f"\n    Erreur: {e}")
            break

        links = data.get("links", [])
        if n_total is None:
            n_total = data.get("numBacklinksTotal", 0)

        for entry in links:
            source = entry.get("source", "")
            try:
                netloc = urlparse(source).netloc.lower().lstrip("www.")
            except Exception:
                continue
            if not netloc or "." not in netloc:
                continue
            try:
                bas = int(entry.get("babbarAuthorityScore", 0) or 0)
            except (ValueError, TypeError):
                bas = 0
            anchor = (entry.get("linkText") or "").strip()

            if netloc not in result:
                result[netloc] = {"bas": bas, "anchors": [anchor] if anchor else []}
            else:
                result[netloc]["bas"] = max(result[netloc]["bas"], bas)
                if anchor and anchor not in result[netloc]["anchors"] and len(result[netloc]["anchors"]) < 3:
                    result[netloc]["anchors"].append(anchor)

        n_links += len(links)
        offset  += 500
        print(f"    [page {page+1}/{max_pages}] {n_links} liens -> {len(result)} domaines", end="\r", flush=True)

        if len(links) < 500:
            break

        time.sleep(PAUSE)

    total_str = f"/{n_total}" if n_total else ""
    print(f"\n  -> {len(result)} domaines ({n_links}{total_str} liens parcourus)\n")
    time.sleep(PAUSE)
    return result


# ── Analyse ───────────────────────────────────────────────────────────────────

def compute_gap(our_domains: set, competitors: dict) -> list:
    """
    Gap = domaines qui linkent >= 1 concurrent mais pas notre site.
    BAS = max observe sur tous les concurrents pour ce domaine.
    """
    all_comp: dict = {}
    for label, domains in competitors.items():
        for domain, info in domains.items():
            if domain not in all_comp:
                all_comp[domain] = {"labels": [], "anchors": info.get("anchors", [])[:5], "bas": info.get("bas", 0)}
            else:
                all_comp[domain]["bas"] = max(all_comp[domain]["bas"], info.get("bas", 0))
            all_comp[domain]["labels"].append(label)

    gap = []
    for domain, info in all_comp.items():
        if domain not in our_domains:
            gap.append({
                "domain":         domain,
                "nb_concurrents": len(info["labels"]),
                "linke":          ", ".join(sorted(set(info["labels"]))),
                "anchors":        " | ".join(a for a in info["anchors"][:3] if a),
                "bas":            info["bas"],
            })

    return gap


def priorite(nb_concurrents: int, bas: int) -> str:
    if nb_concurrents >= 3:
        return "HAUTE"
    elif nb_concurrents >= 2 or bas >= 30:
        return "MOYENNE"
    else:
        return "BASSE"


# ── Export ────────────────────────────────────────────────────────────────────

def generate_excel(gap_comp: list, gap_arm: list, stats: dict, output_path: Path):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("openpyxl manquant -- pip install openpyxl")
        return

    FILL = {
        "HAUTE":   PatternFill("solid", fgColor="D1FAE5"),
        "MOYENNE": PatternFill("solid", fgColor="FEF3C7"),
        "BASSE":   PatternFill("solid", fgColor="FEE2E2"),
    }
    HDR_FILL = PatternFill("solid", fgColor="1A3A5C")
    HDR_FONT = Font(bold=True, color="FFFFFF")

    def write_sheet(ws, gap, title):
        ws.title = title
        headers = [
            "Domaine referent", "Nb concurrents", "BAS",
            "Linkent", "Ancres observees", "Priorite outreach",
        ]
        widths = [42, 16, 10, 55, 60, 18]

        for col, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font, c.fill = HDR_FONT, HDR_FILL
            c.alignment = Alignment(horizontal="center")
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

        for row_i, item in enumerate(gap, 2):
            p = priorite(item["nb_concurrents"], item.get("bas", 0))
            vals = [
                item["domain"],
                item["nb_concurrents"],
                item.get("bas") or "",
                item["linke"],
                item.get("anchors", ""),
                p,
            ]
            for col_i, val in enumerate(vals, 1):
                c = ws.cell(row=row_i, column=col_i, value=val)
                c.fill = FILL[p]

    wb = openpyxl.Workbook()
    ws1 = wb.active
    write_sheet(ws1, gap_comp, "Gap - Comparateurs")

    if gap_arm:
        ws2 = wb.create_sheet("Gap - Armateurs")
        write_sheet(ws2, gap_arm, "Gap - Armateurs")

    ws3 = wb.create_sheet("Stats")
    ws3.append(["Site", "Domaines referents distincts"])
    ws3.cell(1, 1).font = ws3.cell(1, 2).font = Font(bold=True)
    for site, count in stats.items():
        ws3.append([site, count])
    ws3.append([])
    ws3.append(["Gap comparateurs", len(gap_comp)])
    if gap_arm:
        ws3.append(["Gap armateurs", len(gap_arm)])
    ws3.column_dimensions["A"].width = 40
    ws3.column_dimensions["B"].width = 30

    wb.save(output_path)
    print(f"Excel  : {output_path}")


def generate_csv(gap: list, output_path: Path):
    with open(output_path, "w", encoding="utf-8-sig") as f:
        f.write("domaine,nb_concurrents,bas,linke,priorite\n")
        for item in gap:
            p   = priorite(item["nb_concurrents"], item.get("bas", 0))
            bas = item.get("bas") or 0
            f.write(
                f'"{item["domain"]}",'
                f'{item["nb_concurrents"]},'
                f'{bas},'
                f'"{item["linke"]}",'
                f'{p}\n'
            )
    print(f"CSV    : {output_path}")


# ── Main ─────────────────────────────────────────────────────────────────────

def parse_competitors_arg(arg: str) -> dict:
    """
    Parse "label:domain,label:domain" ou "domain,domain" en {label: domain}.
    Si pas de label, utilise le domaine sans www. comme label.
    """
    result = {}
    for part in arg.split(","):
        part = part.strip()
        if not part:
            continue
        if ":" in part:
            label, domain = part.split(":", 1)
            result[label.strip()] = domain.strip()
        else:
            label = part.lstrip("www.").split(".")[0]
            result[label] = part
    return result


def domain_slug(host: str) -> str:
    import re
    return re.sub(r"[^a-z0-9]+", "-", host.lower().replace("www.", "")).strip("-")


def main():
    parser = argparse.ArgumentParser(description="Backlink Gap Analysis -- Babbar API")
    parser.add_argument("--url", default="www.abcroisiere.com",
                        help="Host cible sans https:// (defaut: www.abcroisiere.com)")
    parser.add_argument("--competitors",
                        help='Concurrents principaux. Format: "label:domain,label:domain" ou "domain,domain". '
                             'Si absent, utilise la liste AB Croisiere par defaut.')
    parser.add_argument("--groups",
                        help='Groupe secondaire (ex: armateurs). Meme format que --competitors. '
                             'Produit une sheet separee dans le Excel.')
    parser.add_argument("--include-armateurs", action="store_true",
                        help="Shortcut AB : inclure MSC, Costa, CroisiEurope comme groupe secondaire")
    parser.add_argument("--enrich", action="store_true",
                        help="Enrichir automatiquement le BAS via domain/overview/main apres l'analyse")
    args = parser.parse_args()

    if not API_KEY:
        print("BABBAR_API_KEY manquante dans tools/.env")
        return

    our_host = args.url.replace("https://", "").replace("http://", "").rstrip("/")

    # Resoudre les listes de concurrents
    if args.competitors:
        comparateurs = parse_competitors_arg(args.competitors)
    else:
        comparateurs = COMPARATEURS

    if args.groups:
        armateurs = parse_competitors_arg(args.groups)
        has_groups = True
    elif args.include_armateurs:
        armateurs = ARMATEURS
        has_groups = True
    else:
        armateurs = {}
        has_groups = False

    print(f"\n=== Backlink Gap -- {our_host} ===")
    print(f"Comparateurs : {', '.join(comparateurs.keys())}")
    if has_groups:
        print(f"Groupe 2     : {', '.join(armateurs.keys())}")
    print()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    stats = {}

    # 1. Notre site
    print("[NOUS]")
    our = fetch_domains_with_bas(our_host, our_host, max_pages=MAX_PAGES_US)
    our_domains = set(our.keys())
    stats[our_host] = len(our_domains)

    # 2. Comparateurs
    print("[COMPARATEURS]")
    comp_domains = {}
    for label, host in comparateurs.items():
        comp_domains[label] = fetch_domains_with_bas(host, label)
        stats[label] = len(comp_domains[label])

    # 3. Groupe secondaire (optionnel)
    arm_domains = {}
    if has_groups:
        print("[GROUPE 2]")
        for label, host in armateurs.items():
            arm_domains[label] = fetch_domains_with_bas(host, label)
            stats[label] = len(arm_domains[label])

    # 4. Gap
    print("=== Calcul des gaps ===")
    gap_comp = compute_gap(our_domains, comp_domains)
    gap_arm  = compute_gap(our_domains, arm_domains) if arm_domains else []

    # 5. Tri final : nb_concurrents desc -> BAS desc -> alpha
    gap_comp.sort(key=lambda x: (-x["nb_concurrents"], -x.get("bas", 0), x["domain"]))
    if gap_arm:
        gap_arm.sort(key=lambda x: (-x["nb_concurrents"], -x.get("bas", 0), x["domain"]))

    print(f"Gap comparateurs : {len(gap_comp)} opportunites")
    if gap_arm:
        print(f"Gap groupe 2     : {len(gap_arm)} opportunites")

    # Apercu top 10
    print("\n--- Top 10 (comparateurs) ---")
    for item in gap_comp[:10]:
        bas = item.get("bas", 0)
        p   = priorite(item["nb_concurrents"], bas)
        print(f"  [BAS {bas:>3}] {item['domain']:<42} x{item['nb_concurrents']} concurrents  [{p}]")

    # 6. Export
    print("\n=== Export ===")
    date_str    = datetime.now().strftime("%d-%m-%y")
    site_slug   = domain_slug(our_host)
    group_slug  = "_avec-groupe2" if has_groups else ""
    excel_path  = OUTPUT_DIR / f"backlink_gap_{site_slug}_{date_str}{group_slug}.xlsx"
    csv_path    = OUTPUT_DIR / f"backlink_gap_{site_slug}_{date_str}{group_slug}.csv"

    generate_excel(gap_comp, gap_arm, stats, excel_path)
    generate_csv(gap_comp, csv_path)

    print(f"\nDone -- {len(gap_comp)} opportunites dans output/backlinks/")

    # 7. Enrichissement BAS (optionnel)
    if args.enrich:
        print("\n=== Enrichissement BAS ===")
        try:
            from tools.backlink_gap_enrich import run_enrich
        except ImportError:
            import importlib.util, sys
            spec = importlib.util.spec_from_file_location(
                "backlink_gap_enrich",
                Path(__file__).parent / "backlink_gap_enrich.py"
            )
            mod = importlib.util.load_from_spec(spec)
            spec.loader.exec_module(mod)
            run_enrich = mod.run_enrich
        run_enrich(excel_path, API_KEY)


if __name__ == "__main__":
    main()
