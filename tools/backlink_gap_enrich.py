#!/usr/bin/env python3
"""
backlink_gap_enrich.py — Enrichissement BAS via domain/overview/main (Babbar API)

Lit le fichier Excel produit par backlink_gap.py, appelle domain/overview/main
pour chaque domaine a BAS manquant, et ecrit une copie enrichie.

Usage:
    python tools/backlink_gap_enrich.py
    python tools/backlink_gap_enrich.py --input output/backlinks/backlink_gap_15-04-26_avec-armateurs.xlsx
"""

import argparse
import json
import time
import urllib.request
import urllib.error
from pathlib import Path

# ── Config ────────────────────────────────────────────────────────────────────

PROJECT_ROOT = Path(__file__).parent.parent

def load_env():
    env_path = PROJECT_ROOT / "tools" / ".env"
    env = {}
    if env_path.exists():
        for line in env_path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip()
    return env

ENV     = load_env()
API_KEY = ENV.get("BABBAR_API_KEY", "")

BASE_URL = "https://www.babbar.tech/api"
PAUSE    = 2.1


# ── API ───────────────────────────────────────────────────────────────────────

def _post(endpoint: str, payload: dict) -> dict:
    url  = f"{BASE_URL}/{endpoint}?api_token={API_KEY}"
    data = json.dumps(payload).encode()
    req  = urllib.request.Request(
        url, data=data,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    for attempt in range(10):
        try:
            resp = urllib.request.urlopen(req, timeout=30)
            remaining = int(resp.headers.get("x-ratelimit-remaining", 99))
            if remaining <= 2:
                print(f"\n    Quota bas ({remaining} restants) -- pause 62s...", flush=True)
                time.sleep(62)
            return json.loads(resp.read())
        except urllib.error.HTTPError as e:
            if e.code == 429:
                print(f"    Rate limit 429 -- attente 62s... (tentative {attempt+1}/10)")
                time.sleep(62)
            else:
                body = e.read().decode("utf-8", errors="replace")[:200]
                raise RuntimeError(f"HTTP {e.code} sur {endpoint}: {body}") from e
        except Exception:
            if attempt == 9:
                raise
            time.sleep(5)
    raise RuntimeError(f"Rate limit persistant sur {endpoint} apres 10 tentatives")


def fetch_domain_bas(domain: str) -> tuple[int, str]:
    """
    Retourne (bas, categorie_principale) pour un domaine.
    bas = 0 si non trouve, categorie = "" si absente.
    """
    try:
        data = _post("domain/overview/main", {"domain": domain})
    except Exception as e:
        print(f"\n    Erreur {domain}: {e}")
        return 0, ""

    # L'API retourne [] pour les domaines inconnus
    if not isinstance(data, dict):
        return 0, ""

    bas = 0
    try:
        bas = int(data.get("babbarAuthorityScore", 0) or 0)
    except (ValueError, TypeError):
        pass

    # Categorie principale (langue fr en priorite)
    categorie = ""
    cats = data.get("categories", {})
    lang_cats = cats.get("fr") or cats.get("en") or []
    if lang_cats:
        top = max(lang_cats, key=lambda x: x.get("score", 0))
        categorie = top.get("topicName", "")

    return bas, categorie


# ── Excel helpers ─────────────────────────────────────────────────────────────

def priorite(nb_concurrents: int, bas: int) -> str:
    if nb_concurrents >= 3:
        return "HAUTE"
    elif nb_concurrents >= 2 or bas >= 30:
        return "MOYENNE"
    else:
        return "BASSE"


def enrich_sheet(ws, domain_cache: dict, total_domains: int, enriched_count: list):
    """
    Met a jour BAS + Priorite + ajoute colonne Categorie dans un sheet.
    domain_cache: {domain: (bas, categorie)} -- partage entre sheets pour eviter doublons.
    enriched_count: [int] -- compteur mutable partage.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    FILL = {
        "HAUTE":   PatternFill("solid", fgColor="D1FAE5"),
        "MOYENNE": PatternFill("solid", fgColor="FEF3C7"),
        "BASSE":   PatternFill("solid", fgColor="FEE2E2"),
    }
    HDR_FILL = PatternFill("solid", fgColor="1A3A5C")
    HDR_FONT = Font(bold=True, color="FFFFFF")

    # Verifier / ajouter colonne Categorie (col 7)
    header_row = [ws.cell(1, c).value for c in range(1, 8)]
    if ws.cell(1, 7).value != "Categorie":
        ws.cell(1, 7).value = "Categorie"
        ws.cell(1, 7).font  = HDR_FONT
        ws.cell(1, 7).fill  = HDR_FILL
        ws.cell(1, 7).alignment = Alignment(horizontal="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(7)].width = 30

    for row in ws.iter_rows(min_row=2):
        domain = row[0].value
        if not domain:
            continue

        current_bas = row[2].value or 0

        if domain not in domain_cache:
            if current_bas:
                # BAS deja present -- pas d'appel API
                domain_cache[domain] = (current_bas, "")
            else:
                # BAS manquant -- appel API
                bas, categorie = fetch_domain_bas(domain)
                domain_cache[domain] = (bas, categorie)
                time.sleep(PAUSE)
                enriched_count[0] += 1
                if enriched_count[0] % 20 == 0:
                    print(f"  [{enriched_count[0]}/{total_domains}] domaines enrichis...", flush=True)

        bas, categorie = domain_cache[domain]

        # Mise a jour BAS si manquant ou si on a mieux
        if not current_bas and bas:
            row[2].value = bas

        # Recalcul priorite
        nb_conc = row[1].value or 0
        final_bas = row[2].value or 0
        p = priorite(nb_conc, final_bas)
        row[5].value = p

        # Categorie
        row[6].value = categorie

        # Re-colorier la ligne
        fill = FILL[p]
        for cell in row[:7]:
            cell.fill = fill


# ── Main ─────────────────────────────────────────────────────────────────────

def run_enrich(input_path: Path, api_key: str):
    """
    Enrichit le fichier Excel backlink_gap produit par backlink_gap.py.
    Appelle domain/overview/main pour les domaines a BAS=0, ajoute la colonne Categorie,
    re-trie par BAS enrichi et ecrit {stem}_enriched.xlsx.
    Peut etre appele depuis backlink_gap.py (flag --enrich) ou en standalone.
    """
    global API_KEY
    API_KEY = api_key

    output_path = input_path.with_name(input_path.stem + "_enriched.xlsx")

    print(f"Input  : {input_path.name}")
    print(f"Output : {output_path.name}")

    import shutil
    shutil.copy2(input_path, output_path)
    print(f"Copie creee.\n")

    import openpyxl
    wb = openpyxl.load_workbook(output_path)

    all_domains = set()
    missing_bas = set()
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith("Gap"):
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    all_domains.add(row[0])
                    if not row[2]:
                        missing_bas.add(row[0])

    print(f"Domaines totaux  : {len(all_domains)}")
    print(f"BAS manquants    : {len(missing_bas)} (seuls ceux-la feront un appel API)")
    duree_min = len(missing_bas) * PAUSE / 60
    print(f"Duree estimee    : ~{duree_min:.0f} min\n")

    domain_cache = {}
    enriched_count = [0]

    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("Gap"):
            continue
        print(f"=== {sheet_name} ===")
        ws = wb[sheet_name]
        enrich_sheet(ws, domain_cache, len(all_domains), enriched_count)
        print(f"  Sheet termine.\n")

    # Re-trier chaque sheet Gap : nb_concurrents desc -> BAS desc -> alpha
    from openpyxl.styles import PatternFill
    FILL = {
        "HAUTE":   PatternFill("solid", fgColor="D1FAE5"),
        "MOYENNE": PatternFill("solid", fgColor="FEF3C7"),
        "BASSE":   PatternFill("solid", fgColor="FEE2E2"),
    }
    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("Gap"):
            continue
        ws = wb[sheet_name]
        rows_data = [[c.value for c in row] for row in ws.iter_rows(min_row=2)]
        rows_data.sort(key=lambda x: (-(x[1] or 0), -(x[2] or 0), (x[0] or "")))
        for r_idx, vals in enumerate(rows_data, 2):
            p = vals[5] or "BASSE"
            fill = FILL.get(p, FILL["BASSE"])
            for c_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.fill = fill

    wb.save(output_path)
    print(f"Enrichissement termine. Fichier : {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Enrichissement BAS -- Babbar domain/overview/main")
    parser.add_argument(
        "--input",
        default=None,
        help="Chemin vers le fichier Excel backlink_gap (defaut: dernier fichier dans output/backlinks/)"
    )
    args = parser.parse_args()

    if not API_KEY:
        print("BABBAR_API_KEY manquante dans tools/.env")
        return

    if args.input:
        input_path = Path(args.input)
    else:
        bl_dir = PROJECT_ROOT / "output" / "backlinks"
        candidates = sorted(bl_dir.glob("backlink_gap_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
        candidates = [p for p in candidates if "_enriched" not in p.name]
        if not candidates:
            print("Aucun fichier backlink_gap_*.xlsx trouve dans output/backlinks/")
            return
        input_path = candidates[0]

    run_enrich(input_path, API_KEY)


if __name__ == "__main__":
    main()
